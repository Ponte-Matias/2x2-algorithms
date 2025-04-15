import streamlit as st
from openpyxl import load_workbook
import random
from algoritmo_inversor import invertir_alg

# Función que obtiene todas las celdas que ocupa verticalmente una combinación de celdas (las que están en la primer fila e indican el subset)
def obtener_rango_combinado_y_valores(hoja, valor_buscado):

    rango_vertical = None

    # Paso 1: Buscar la celda combinada que contiene el valor buscado
    for rango in hoja.merged_cells.ranges:
        celda_inicial = hoja[rango.coord.split(":")[0]]
        if celda_inicial.value == valor_buscado:
            rango_vertical = rango
            break

    if rango_vertical is None:
        raise ValueError(f"No se encontró una celda combinada con el valor '{valor_buscado}'")

    # Paso 2: Extraer coordenadas del rango (ej: A3:A8)
    start_cell, end_cell = str(rango_vertical).split(":")
    start_col = hoja[start_cell].column_letter
    start_row = hoja[start_cell].row
    end_row = hoja[end_cell].row

    # Paso 3: Recorrer columnas B en adelante, en esas filas
    datos = []

    for fila in range(start_row, end_row + 1):
        for columna in range(2, hoja.max_column + 1):  # columna 2 = B
            celda = hoja.cell(row=fila, column=columna)
            if celda.value is not None:
                datos.append({
                    "fila": fila,
                    "columna": celda.column_letter,
                    "algoritmo": celda.value
                })

    return datos

# Cargar wb de forma estática ya que nunca cambia
@st.cache_resource
def cargar_workbook(nombre_archivo):
    wb = load_workbook(nombre_archivo, data_only=True)
    return wb

# Ruta al archivo Excel
archivo = "algs.xlsx"
# Usar el workbook cargado
wb = cargar_workbook(archivo)

st.title("2x2 Set-Up Algorithms")   # Titulo
st.markdown("---")  # Línea divisoria

metodo = st.sidebar.selectbox("Method: ", wb.sheetnames)    # Metodos/hojas posibles
# Escoger el método, acceder a la hoja correspondiente
# Guardar la hoja si cambió el método
cambio = 0      # Para verificar mas adelante
if "metodo_actual" not in st.session_state or st.session_state.metodo_actual != metodo:     # Entra la primera vez y cuando se cambia de hoja
    st.session_state.metodo_actual = metodo
    st.session_state.hoja_actual = wb[metodo]
    cambio = 1      # Si se cambia de método/hoja, esta este verificador

# Usar la hoja desde session_state
hoja = st.session_state.hoja_actual

# Recorrer la columna A (columna 1), ignorando celdas vacías
subsets = []    # Lista con los posibles subsets del metodo
for fila in hoja.iter_rows(min_col=1, max_col=1):
    celda = fila[0]
    if celda.value is not None:
        subsets.append(celda.value) 

subsets.pop(0)  # Elimino el elemento que contiene el nombre del método
subsets_a_mostrar = []
for subset in subsets:
    if st.toggle(subset):
        subsets_a_mostrar.append(subset)

# Botón de loop
loop_opcion = st.checkbox("Loop mode")
if "loop_opcion" not in st.session_state:
    st.session_state.loop_opcion = loop_opcion      # Para chequear más adelante si es tildado luego de estar destildado

if "subsets_a_mostrar" not in st.session_state:     # Inicializa el estado la primera vez
    st.session_state.subsets_a_mostrar = subsets_a_mostrar

# BLOQUE PRINCIPAL
recalcular_todo = 0         # Verificador para el boton de loop
if len(subsets_a_mostrar) != 0:     # Por si esta vacio (al inicio mas que nada)
    if st.session_state.subsets_a_mostrar != subsets_a_mostrar or cambio == 1:      # No recalcular todo si no se cambia de hoja o los subsets a mostrar
        st.session_state.subsets_a_mostrar = subsets_a_mostrar      # Actualizar estado
        recalcular_todo = 1         # Verificador para el boton de loop
        # Leer cada fila por separado, con el vector subsets_a_mostrar
        vector_total = []
        for fila in subsets_a_mostrar:
            valores = obtener_rango_combinado_y_valores(hoja, fila)  # valores: lista con diccionarios
            # total_casos obtiene todas las columnas en donde tiene algoritmos (tupla para no repetir las columnas)
            total_casos = set({})   # Indica qué columnas tienen algs (la cantidad de casos del subset)
            for dato in valores:
                total_casos.add(dato['columna'])
            # Pasarlo a lista y ordenarlo alfabeticamente
            total_casos = list(total_casos)
            total_casos.sort()

            # Leer cada columna por separado, con el vector total_casos
            for columna in total_casos:
                # Quedarse con un solo algoritmo para un caso en especifico
                algs_un_caso = []
                for dato in valores:
                    if dato['columna'] == columna:
                        algs_un_caso.append(dato['algoritmo'])
                numero = random.randint(0, len(algs_un_caso)-1)
                vector_total.append(algs_un_caso[numero])
        # Limpiar vector_total de caracteres nulos y espacios
        vector_total = list(filter(None, vector_total))
        vector_total = [x for x in vector_total if x != ' ']

        # Filtrar los elementos que tienen contenido real, si hay algo vacio se va
        vector_total = [x for x in vector_total if x.strip() != '']

        # Invertir todos los algoritmos, usando la funcion en el otro script
        invertido_total = []
        for algoritmo in vector_total:
            invertido_total.append(invertir_alg(algoritmo))
        st.session_state.invertido_total = invertido_total      # Guardar estado
    
    invertido_total = st.session_state.invertido_total          # Cargar estado para cuando no se recalcula (cuando no se entra en el if)
    
    # Lógica de loop y random
    if loop_opcion == False:
        numero = random.randint(0, len(invertido_total)-1)
    else:
        # Inicializar la primera vez que se usa el boton de loop o si se recalcula todo se reinicia o si se llega al ultimo alg luego se reinicia o si se tilda el boton loop luego de estar destildado
        if "numero" not in st.session_state or recalcular_todo == 1:
            st.session_state.numero = -1        # Para que en la siguiente iteracion (al apretar Next alg) se vaya a 0
            numero = st.session_state.numero
        elif st.session_state.numero == len(invertido_total)-1:
            st.session_state.numero = 0        
            numero = st.session_state.numero
        elif st.session_state.loop_opcion != loop_opcion:
            st.session_state.loop_opcion = loop_opcion
            st.session_state.numero = 0        
            numero = st.session_state.numero
        else:
            st.session_state.numero += 1
            numero = st.session_state.numero

    # Inicializar la variable en la sesión si no existe
    if "alg_random" not in st.session_state:
        st.session_state.alg_random = invertido_total[numero]

    # Botón para cambiar de algoritmo
    if st.button("Next alg"):
        st.session_state.alg_random = invertido_total[numero]

    # Mostrar el algoritmo seleccionado
    st.subheader(st.session_state.alg_random)

# Aclaraciones y demás
st.markdown("---")  # Línea divisoria
st.markdown(
    """
    **Notes**:
    - This does not generate any algorithm, it just takes the ones shown in this [drive](https://docs.google.com/spreadsheets/d/1OFXakCV85Mp2zsQBXMxiMX9a506JeAcLnUXZr8FgXAY/) and then reverse them.
    - If you do not press the Next alg button, the shown algorithm is not going to change, even if you change the method.
    - When in loop mode, you would reset the loop if you add or delete a subset
    **Tip**: if you are in PC, you can click Next alg and then press the enter key or the space bar to continue changing the algs.
    
    Found any bug? Contact me: [mail](matiasponte20@gmail.com)
    """,
    unsafe_allow_html=True
)
