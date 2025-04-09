from openpyxl import load_workbook
import random
from algoritmo_inversor import invertir_alg

# Función que obtiene todas las celdas que ocupa verticalmente una combinación de celdas
# (las que están en la primer fila e indican el subset)
def obtener_rango_combinado_y_valores(hoja, valor_buscado):

    rango_vertical = None

    # Paso 1: Buscar la celda combinada que contiene el valor buscado
    for rango in hoja.merged_cells.ranges:
        celda_inicial = hoja[rango.coord.split(":")[0]]
        if celda_inicial.value == valor_buscado:
            rango_vertical = rango
            break

    if rango_vertical is None:
        raise ValueError(f"No se encontró una celda combinada con el valor '{valor_buscado}'")

    # Paso 2: Extraer coordenadas del rango (ej: A3:A8)
    start_cell, end_cell = str(rango_vertical).split(":")
    start_col = hoja[start_cell].column_letter
    start_row = hoja[start_cell].row
    end_row = hoja[end_cell].row

    print(f"'{valor_buscado}' ocupa de {start_cell} a {end_cell} (filas {start_row} a {end_row})")

    # Paso 3: Recorrer columnas B en adelante, en esas filas
    datos = []

    for fila in range(start_row, end_row + 1):
        for columna in range(2, hoja.max_column + 1):  # columna 2 = B
            celda = hoja.cell(row=fila, column=columna)
            if celda.value is not None:
                datos.append({
                    "fila": fila,
                    "columna": celda.column_letter,
                    "algoritmo": celda.value
                })

    return datos

# Ruta al archivo Excel
archivo = "algs.xlsx"

# Cargar el archivo
wb = load_workbook(archivo, data_only=True)

# Escoger el método
# IMPORTANTE: borrar las hojas que no sean de algoritmos: Home, PBL Angles, Copy of LS-4, etc
# Inclusive las ocultas (click derecho abajo, mostrar y luego borrar)
i=1
for elemento in wb.sheetnames:
    print(i, " - ", elemento)
    i+=1
opcion = int(input("Metodo a buscar: "))
metodo = wb.sheetnames[opcion-1]
# Acceder a la hoja correspondiente
if metodo in wb.sheetnames:
    hoja = wb[metodo]
else:
    raise ValueError("La hoja 'TCLL+' no existe en el archivo.")

# Recorrer la columna A (columna 1), ignorando celdas vacías
subsets = []    # Lista con los posibles subsets del metodo
for fila in hoja.iter_rows(min_col=1, max_col=1):
    celda = fila[0]
    if celda.value is not None:
        print(celda.value)
        subsets.append(celda.value) 

print("---------------")
subsets.pop(0)  # Elimino el elemento que contiene el nombre del método
print(subsets)
""" Esto se puede borrar si no me equivoco, 99% seguro
elemento = subsets[1]   # Hammer en este caso

valores = obtener_rango_combinado_y_valores(hoja, elemento)  # valores: lista con diccionarios

# Imprime la tabla
for dato in valores:
    print(f"Fila {dato['fila']} | Columna {dato['columna']} | Algoritmo: {dato['algoritmo']}")

print("---------------")
print(valores)

# Esto obtiene todas las columnas en donde tiene algoritmos (tupla para no repetir las columnas)
total_casos = set({})   # Indica qué columnas tienen algs (la cantidad de casos del subset)
for dato in valores:
    total_casos.add(dato['columna'])
# Pasarlo a lista y ordenarlo alfabeticamente
total_casos = list(total_casos)
total_casos.sort()
print(total_casos)
"""
# BLOQUE PRINCIPAL
# Leer cada fila por separado, con el vector subsets
vector_total = []
for fila in subsets:
    valores = obtener_rango_combinado_y_valores(hoja, fila)  # valores: lista con diccionarios
    # total_casos obtiene todas las columnas en donde tiene algoritmos (tupla para no repetir las columnas)
    total_casos = set({})   # Indica qué columnas tienen algs (la cantidad de casos del subset)
    for dato in valores:
        total_casos.add(dato['columna'])
    # Pasarlo a lista y ordenarlo alfabeticamente
    total_casos = list(total_casos)
    total_casos.sort()

    # Leer cada columna por separado, con el vector total_casos
    for columna in total_casos:
        # Quedarse con un solo algoritmo para un caso en especifico
        algs_un_caso = []
        for dato in valores:
            if dato['columna'] == columna:
                #print(dato['algoritmo'])
                algs_un_caso.append(dato['algoritmo'])
        numero = random.randint(0, len(algs_un_caso)-1)
        #print(numero)
        vector_total.append(algs_un_caso[numero])
# Limpiar vector_total de caracteres nulos y espacios
vector_total = list(filter(None, vector_total))
vector_total = [x for x in vector_total if x != ' ']

print(vector_total)
print(len(vector_total))

# Invertir todos los algoritmos, usando la funcion en el otro script
invertido_total = []
for algoritmo in vector_total:
    invertido_total.append(invertir_alg(algoritmo))


while True:
    try:
        numero = random.randint(0, len(invertido_total)-1)
        print(invertido_total[numero])
        espera = input("Presione enter para continuar, Ctrl + C para salir")
    except KeyboardInterrupt:
        print("\nSaliendo del programa...")
        exit(0)
